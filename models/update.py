# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import requests
from sqlalchemy.dialects.postgresql import insert

import datetime
import pathlib
import json

from models.database import engine
from models.models import corona_data


def update_database():
    url = (
        # 大阪府のコロナ特設サイトのgithubのjsonファイルのリンク。
        "https://github.com/codeforosaka/covid19/raw/"
        "master/data/data.json"
    )
    print("Downloading...")

    download_file = requests.get(url)
    """
    # ファイルを保存するパスの指定と存在確認
    path = pathlib.Path(__file__).joinpath(
        "..", "..", "tmp"
    ).resolve()
    if path.exists() is False:
        path.mkdir()
    path = path.joinpath("corona_data.json")
    with path.open(mode="wb") as f:
        f.write(download_file.content)
    """
    jsonb = download_file.content
    json_data = json.loads(jsonb)
    # print(len(json_data["patients"]["data"]))
    r_list = []
    for i in range(len(json_data["patients"]["data"])):
        r_tpl = ()
        r_tpl += json_data["patients"]["data"][i][0]
        r_tpl += json_data["patients"]["data"][i][1]
        r_tpl += json_data["patients"]["data"][i][2]
        r_tpl += json_data["patients"]["data"][i][3]
        r_tpl += json_data["patients"]["data"][i][4]
        r_tpl += json_data["patients"]["data"][i][5]
        r_tpl += json_data["patients"]["data"][i][6]
        r_tpl += json_data["patients"]["data"][i][7]
        r_list.append(r_tpl)

    """
    wb = load_workbook(filename=str(path), data_only=True)
    ws = wb.worksheets[1]
    r_list = []
    # +1 する理由 : range関数はrange(start, stop)で
    # start=< i <stopの連番を作成する。
    # 追加のメモ。max_row, columnはデータが無くとも書式が設定されていれば
    # 反応する
    print("loading file...")
    for r in range(3, ws.max_row):
        #print(f"{r}:")
        r_tpl = ()
        tmp = ws.cell(r, 1).value
        if tmp is None:
            break
        for c in range(1, 9):
            #print(c)
            value = ws.cell(r, c).value
            if c == 2:
                value = datetime.date(
                    year=1900, month=1, day=1
                ) + datetime.timedelta(days=value - 2)
            if isinstance(value, datetime.datetime):
                value = value.date()
            r_tpl += (value, )
        r_list.append(r_tpl)
    """
    insert_stmt = insert(corona_data)
    set_ = dict(
        index=insert_stmt.excluded.index,
        publish_d=insert_stmt.excluded.publish_d,
        age=insert_stmt.excluded.age,
        gender=insert_stmt.excluded.gender,
        place=insert_stmt.excluded.place,
        date_of_onset=insert_stmt.excluded.date_of_onset,
        symptoms=insert_stmt.excluded.symptoms,
        hospitalization=insert_stmt.excluded.hospitalization
    )
    insert_stmt = insert_stmt.on_conflict_do_update(
        index_elements=["index"], set_=set_
    )
    print("insert now...")
    with engine.connect() as conn:
        values = []
        for i in range(0, len(r_list)):
            data = {}
            data["index"] = r_list[i][0]
            data["publish_d"] = r_list[i][1]
            data["age"] = r_list[i][2]
            data["gender"] = r_list[i][3]
            data["place"] = r_list[i][4]
            data["date_of_onset"] = r_list[i][5]
            data["symptoms"] = r_list[i][6]
            data["hospitalization"] = r_list[i][7]
            values.append(data)
        conn.execute(insert_stmt, values)
    print("successfully!!")
